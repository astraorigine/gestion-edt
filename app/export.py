# app/export.py
# ─────────────────────────────────────────
# Export PDF et Excel des emplois du temps
# Projet : Gestion EDT — Méthode XP
# Itération 4 — GEDT-08
# ─────────────────────────────────────────

import os
from datetime import datetime
from app.models import (
    EmploiDuTemps, Seance,
    Matiere, Enseignant,
    Parcours, Semestre
)


def _get_donnees(db, edt_id):
    """
    Récupère toutes les données
    nécessaires pour l'export.
    """
    edt = db.query(EmploiDuTemps).filter_by(
        id=edt_id
    ).first()

    if not edt:
        raise ValueError(
            f"Emploi du temps #{edt_id} "
            f"introuvable."
        )

    p = db.query(Parcours).filter_by(
        id=edt.parcours_id
    ).first()
    s = db.query(Semestre).filter_by(
        id=edt.semestre_id
    ).first()

    seances = db.query(Seance).filter_by(
        emploi_du_temps_id=edt_id
    ).order_by(
        Seance.jour, Seance.heure_debut
    ).all()

    seances_enrichies = []
    for seance in seances:
        mat = db.query(Matiere).filter_by(
            id=seance.matiere_id
        ).first()
        ens = db.query(Enseignant).filter_by(
            id=seance.enseignant_id
        ).first()
        seances_enrichies.append({
            "jour"      : seance.jour,
            "heure"     : seance.heure_debut,
            "duree"     : seance.duree,
            "matiere"   : mat.nom if mat else "?",
            "couleur"   : mat.type.couleur if mat else "noir",
            "enseignant": (
                f"{ens.grade} {ens.nom}"
                if ens else "Non assigné"
            )
        })

    return {
        "edt"     : edt,
        "parcours": p.nom if p else "?",
        "semestre": s.numero if s else "?",
        "seances" : seances_enrichies
    }


def generer_excel(db, edt_id):
    """
    Génère un fichier Excel de l'emploi
    du temps et retourne son chemin.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill,
            Alignment, Border, Side
        )
    except ImportError:
        raise ValueError(
            "openpyxl non installé. "
            "Lancez : pip install openpyxl"
        )

    donnees = _get_donnees(db, edt_id)
    edt     = donnees["edt"]

    # Couleurs de remplissage selon colorimétrie
    couleurs_excel = {
        "bleu" : "BBDEFB",
        "jaune": "FFF9C4",
        "noir" : "F5F5F5",
        "rouge": "FFCDD2",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ET_{edt_id}"

    # Titre
    ws.merge_cells("A1:E1")
    ws["A1"] = (
        f"EMPLOI DU TEMPS — "
        f"{edt.session.upper()} | "
        f"{donnees['parcours']} — "
        f"Semestre {donnees['semestre']}"
    )
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(
        horizontal="center"
    )
    ws["A1"].fill = PatternFill(
        fill_type="solid", fgColor="1A1F3B"
    )
    ws["A1"].font = Font(
        bold=True, size=13, color="FFFFFF"
    )

    # Sous-titre
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Publié le : {edt.date_publication} | "
        f"Début : {edt.date_debut}"
    )
    ws["A2"].alignment = Alignment(
        horizontal="center"
    )
    ws["A2"].fill = PatternFill(
        fill_type="solid", fgColor="252B50"
    )
    ws["A2"].font = Font(
        size=11, color="D4E4FF"
    )

    # En-têtes colonnes
    entetes = [
        "Jour", "Horaire",
        "Matière", "Type", "Enseignant"
    ]
    for col, ent in enumerate(entetes, 1):
        cell = ws.cell(row=3, column=col, value=ent)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill(
            fill_type="solid", fgColor="2C3E50"
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

    # Données
    thin = Side(style="thin", color="CCCCCC")
    border = Border(
        left=thin, right=thin,
        top=thin, bottom=thin
    )

    for i, s in enumerate(
        donnees["seances"], start=4
    ):
        heure_str = str(s["heure"])[:5]
        valeurs   = [
            s["jour"],
            f"{heure_str} ({s['duree']}h)",
            s["matiere"],
            s["couleur"].capitalize(),
            s["enseignant"]
        ]
        fill_color = couleurs_excel.get(
            s["couleur"], "FFFFFF"
        )
        for col, val in enumerate(valeurs, 1):
            cell = ws.cell(
                row=i, column=col, value=val
            )
            cell.fill   = PatternFill(
                fill_type="solid",
                fgColor=fill_color
            )
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

    # Largeur colonnes
    for col, width in zip(
        ["A","B","C","D","E"],
        [14, 16, 30, 14, 28]
    ):
        ws.column_dimensions[col].width = width

    # Sauvegarder
    dossier = os.path.join(
        os.path.dirname(
            os.path.dirname(
                os.path.abspath(__file__)
            )
        ),
        "exports"
    )
    os.makedirs(dossier, exist_ok=True)
    chemin = os.path.join(
        dossier, f"ET_{edt_id}.xlsx"
    )
    wb.save(chemin)
    return chemin


def generer_pdf(db, edt_id):
    """
    Génère un fichier PDF de l'emploi
    du temps et retourne son chemin.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table,
            TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import (
            getSampleStyleSheet, ParagraphStyle
        )
    except ImportError:
        raise ValueError(
            "reportlab non installé. "
            "Lancez : pip install reportlab"
        )

    donnees = _get_donnees(db, edt_id)
    edt     = donnees["edt"]

    # Couleurs ReportLab
    couleurs_pdf = {
        "bleu" : colors.HexColor("#BBDEFB"),
        "jaune": colors.HexColor("#FFF9C4"),
        "noir" : colors.HexColor("#F5F5F5"),
        "rouge": colors.HexColor("#FFCDD2"),
    }

    dossier = os.path.join(
        os.path.dirname(
            os.path.dirname(
                os.path.abspath(__file__)
            )
        ),
        "exports"
    )
    os.makedirs(dossier, exist_ok=True)
    chemin = os.path.join(
        dossier, f"ET_{edt_id}.pdf"
    )

    doc    = SimpleDocTemplate(chemin, pagesize=A4)
    styles = getSampleStyleSheet()
    story  = []

    # Titre
    titre_style = ParagraphStyle(
        "titre",
        fontSize  = 14,
        fontName  = "Helvetica-Bold",
        textColor = colors.HexColor("#1A1F3B"),
        spaceAfter= 6
    )
    story.append(Paragraph(
        f"EMPLOI DU TEMPS — "
        f"{edt.session.upper()}",
        titre_style
    ))
    story.append(Paragraph(
        f"{donnees['parcours']} | "
        f"Semestre {donnees['semestre']} | "
        f"Début : {edt.date_debut}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.4*cm))

    # Tableau
    data = [[
        "Jour", "Horaire",
        "Matière", "Type", "Enseignant"
    ]]

    styles_tableau = [
        ("BACKGROUND", (0,0), (-1,0),
         colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",  (0,0), (-1,0),
         colors.white),
        ("FONTNAME",   (0,0), (-1,0),
         "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("GRID",       (0,0), (-1,-1),
         0.5, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.white, colors.HexColor("#F8F9FF")]),
    ]

    for i, s in enumerate(
        donnees["seances"], start=1
    ):
        heure_str = str(s["heure"])[:5]
        data.append([
            s["jour"],
            f"{heure_str} ({s['duree']}h)",
            s["matiere"],
            s["couleur"].capitalize(),
            s["enseignant"]
        ])
        bg = couleurs_pdf.get(
            s["couleur"],
            colors.white
        )
        styles_tableau.append((
            "BACKGROUND",
            (0, i), (-1, i), bg
        ))

    table = Table(
        data,
        colWidths=[
            2.5*cm, 3*cm, 5.5*cm,
            2.5*cm, 5*cm
        ]
    )
    table.setStyle(TableStyle(styles_tableau))
    story.append(table)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Document généré le "
        f"{datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        styles["Normal"]
    ))

    doc.build(story)
    return chemin